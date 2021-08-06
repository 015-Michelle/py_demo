import datetime
import openpyxl
import json


def read_json():
    with open('id.json', 'r', encoding='UTF-8') as json_file:
        translation = json.load(json_file)
        return translation


def write(first, second):
    bg = openpyxl.load_workbook('GCPCard.xlsx')  # 应先将excel文件放入到工作目录下
    sheet = bg["App"]  # “App”表示将数据写入到excel文件的App下
    # sheet.cell(0, 1, "key")
    # sheet.cell(0, 2, "英文(en)")
    for i in range(1, len(first) + 1):
        sheet.cell(i, 1, first[i - 1])
        sheet.cell(i, 2, second[i - 1])  # sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列

    bg.save('GCPCard.xlsx')  # 对文件进行保存


print('processing...')
start = datetime.datetime.now()
result = read_json()
keys = []
values = []
for key in result:
    keys.append(key)
    values.append(result[key])
print(values)
write(keys, values)
print('done...')
end = datetime.datetime.now()
print('cost ', end - start, 'ms')
