import datetime

import openpyxl
import json


dir = 'E:\\DeskTop\\'


def read_json():
    with open('translate.json', 'r', encoding='UTF-8') as json_file:
        translation = json.load(json_file)
        return translation


def handle(json: dict):
    wb = openpyxl.load_workbook('GCPCard.xlsx')
    sheet = wb['App']
    # row2 = [item.value for item in list(sheet.rows)[1]]
    # print('第二行值', row2)
    # col1 = [item.value for item in list(sheet.columns)[0]]
    # print('第一列的值', col1)
    # cell_2_3 = sheet.cell(row=2, column=3).value
    # print(cell_2_3)
    # print('第二行3列单元格值', cell_2_3)
    # max_row = sheet.max_row
    # print('行数', max_row)

    for i in range(1, max_row):
        translated = sheet.cell(row=i, column=6).value

        if translated is not None and translated.strip() in json.values():
            key = list(json.keys())[list(json.values()).index(translated.strip())]
            sheet.cell(row=i, column=2, value=key)

    wb.save('GcpCard.xlsx')


if __name__ == '__main__':
    print('processing...')
    start = datetime.datetime.now()
    json = read_json()
    handle(json)
    print('done...')
    end = datetime.datetime.now()
    print('cost ', end - start, 'ms')

